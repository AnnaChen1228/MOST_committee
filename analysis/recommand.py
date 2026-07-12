import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from save_commitee_data import filiter_school 

# --- 1. 利益迴避核心判定模組 (核心修改：獨立出來方便日後擴充) ---
def check_conflict_status(cand_name, cand_school, apply_advisors, app_info, all_applicants_set, blacklist):
    """
    判定委員與計畫之間的衝突狀態。
    回傳值：(衝突類型字串, 建議顏色代碼)
    """
    app_name = app_info.get('manager', '')
    app_school = app_info.get('school', '')
    co_managers = app_info.get('co_managers', [])
    co_schools = app_info.get('co_schools', [])

    # 1. 黑名單 (優先權最高)
    if cand_name in blacklist:
        return "黑名單", "gray"
    
    # 2. 本人迴避
    if cand_name == app_name:
        return "本人", "green"
    
    # 3. 申請者迴避 (本次所有計畫的申請人)
    if cand_name in all_applicants_set:
        return "為本次計畫申請者", "green"
    
    # 4. 共同主持人迴避
    if cand_name in co_managers:
        return "是推薦人", "blue"
    
    # 5. 主持人同校迴避
    if app_school and cand_school and app_school == cand_school:
        return "與申請人同校", "red"
    
    # 6. 共同主持人同校迴避
    if cand_school and cand_school in co_schools:
        return "與推薦人同校", "blue"
    
    # 7. 指導教授迴避
    if cand_name in apply_advisors:
        return "為申請者博士指導教授", "red"

    return None, None

# --- 2. 讀取與計算邏輯 (維持不變) ---
def load_apply_data(file_path, pages):
    apply_dicts = {} 
    for page in pages:
        try:
            df = pd.read_excel(file_path, sheet_name=page, dtype=str).fillna("")
            for _, row in df.iterrows():
                title = str(row.get('研習主題') or row.get('計畫名稱') or "").strip()
                no =  str(row.get('計畫編號')).strip()
                if not title: continue
                school, dept = filiter_school(row.get('申請機構'))
                
                # 處理共同主持人
                co_m, co_s = [], []
                co_manager_raw = row.get('共同主持人') 
                
                if pd.notna(co_manager_raw) and str(co_manager_raw).strip():
                    for item in str(co_manager_raw).split(';'):
                        match = re.match(r'^(.*?)[(（](.*)[)）]', item.strip())
                        if match:
                            name, s_raw = match.group(1).strip(), match.group(2).strip()
                            s_clean, _ = filiter_school(s_raw)
                            co_m.append(name); co_s.append(s_clean)
                        else:
                            co_m.append(item.strip())
                apply_dicts[title] = {
                    'no': no,
                    'manager': str(row['主持人']).strip(),
                    'managerSchool': school,
                    'coManger': co_m,
                    'coSchool': co_s
                }
        except Exception as e:
            print(f"跳過頁面 {page}: {e}")
            continue
    return apply_dicts

def load_score_data(file_path, pages):
    page_data = {}
    for page in pages:
        try:
            df = pd.read_excel(file_path, sheet_name=page, dtype=str).fillna("")
            page_data[page] = df[['project', 'manager', 'similarity_score', 'recommended_manager']]
        except Exception as e:
            print(f"跳過頁面 {page}: {e}")
            continue
    return page_data

def calculate_score(data):
    weights = {'title': 1, 'keyword': 1, 'application_directions': 3, 'problems_to_solve': 3, 'goals_to_achieve': 1, 'methods_to_solve': 1}
    total_w = sum(weights.values())
    results = {}
    for proj, pages_data in data.items():
        m_scores = {}
        app_m = ""
        for pg, rows in pages_data.items():
            w = weights.get(pg, 0)
            for r in rows:
                if not app_m: app_m = r.get('manager', '')
                rec = r.get('recommended_manager')
                if rec:
                    try: s = float(r.get('similarity_score', 0))
                    except: s = 0
                    m_scores[rec] = m_scores.get(rec, 0) + (s * w)
        cand_list = sorted([{'name': m, 'score': s/total_w} for m, s in m_scores.items()], key=lambda x: x['score'], reverse=True)
        results[proj] = {'manager': app_m, 'school': "", 'candidates': cand_list}
    return results

# --- 3. 存檔並上色 (調用模組化判定) ---
def save_results_with_highlight(results, output_filename, reviewer_school_map, reviewer_advisor_map, blacklist, all_applicants_set):
    excel_rows = []
    # sorted_projects = sorted(results.keys())
    
    for project in results.keys():
        info = results[project]
        # 基礎資料
        row = {"計畫編號": info['no'], "計畫名稱": project, '主持人': info['manager'], "申請機構": info['school']}
        
        reasons_list = [] # 用來收集這一案前 10 名的衝突理由
        
        apply_advisors = reviewer_advisor_map.get(info['manager'], [])
        # 處理前 10 位候選人
        for i, cand in enumerate(info.get('candidates', [])[:10]):
            cand_name = cand['name']
            cand_school = reviewer_school_map.get(cand_name, "")
            row[f"推薦委員 {i+1}"] = cand_name
            row[f"相似度分數 {i+1}"] = cand['score']
            row[f"推薦委員學校 {i+1}"] = cand_school
            # row[f"推薦委員名稱學校 {i+1}"] = f"{cand['name']}({cand.get('school', '')})"
            # --- 新增：在寫入資料時就先判定理由 ---
            reason_type, _ = check_conflict_status(cand_name, cand_school, apply_advisors, info, all_applicants_set, blacklist)
            if reason_type:
                reasons_list.append(f"{cand_name}:{reason_type}")
        
        # 將所有理由串接起來，放在最後一欄
        if reasons_list:
            row["過濾原因"] = "[" + ";".join(reasons_list) + ";]"
        else:
            row["過濾原因"] = "[]"

        excel_rows.append(row)

    # 存成 Excel
    pd.DataFrame(excel_rows).to_excel(output_filename, index=False)

    # --- 後續標色處理 ---
    wb = load_workbook(output_filename)
    ws = wb.active
    color_map = {
        "gray": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "green": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "blue": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "red": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    }

    for i, project in enumerate(results):
        row_idx = i + 2
        info = results[project]
        for k, cand in enumerate(info.get('candidates', [])[:10]):
            cand_name = cand['name']
            cand_school = reviewer_school_map.get(cand_name, "")
            
            # 再次判定顏色（確保與理由同步）
            _, color_key = check_conflict_status(cand_name, cand_school, apply_advisors, info, all_applicants_set, blacklist)
            
            if color_key:
                # 計算欄位位置：Project(1), Manager(2), School(3) 後開始，每 3 欄一組
                base_col = 5 + (k * 3)
                for c in range(3): 
                    ws.cell(row=row_idx, column=base_col+c).fill = color_map[color_key]
    
    wb.save(output_filename)

# --- 4. 存檔乾淨名單 (含過濾理由) ---
def save_results_clean(results, output_filename):
    excel_rows = []
    for project in results.keys():
        info = results[project]
        row = {"計畫編號": info['no'], "計畫名稱": project, "主持人": info['manager'], "申請機構": info['school'], "過濾原因": info.get('filter_reason_str', "")}
        
        for i, cand in enumerate(info.get('final_candidates', [])[:10]):
            row[f"推薦委員 {i+1}"] = cand['name']; row[f"相似度分數 {i+1}"] = cand['score']; row[f"推薦委員學校 {i+1}"] = cand.get('school', '') 
            # row[f"推薦委員名稱學校 {i+1}"] = f"{cand['name']}({cand.get('school', '')})"
        excel_rows.append(row)
    pd.DataFrame(excel_rows).to_excel(output_filename, index=False)

def group_by_project_dict(datas, pages):
    results = {}
    for page in pages:
        if page not in datas: continue
        for row in datas[page].to_dict('records'):
            p = row.get('project')
            if p:
                if p not in results: results[p] = {}
                if page not in results[p]: results[p][page] = []
                results[p][page].append(row)
    return results

def check_path(path):
    if not os.path.exists(path):
        print(f"找不到檔案: {path}")

# --- 5. 主程式 ---
def main():
    apply_path = 'data/research_proj/115計算機學門審查/瑞典/apply_project_with_abstract(計畫書資料).xlsx'
    check_path(apply_path)
    input_score_file = 'data/research_proj/115計算機學門審查/瑞典/result_score(計畫書資料_research).xlsx'
    check_path(input_score_file)
    committee_uni_file = 'data/RDF_database/committee_all_education_with_advisor.xlsx'
    check_path(committee_uni_file)
    blacklist_path = 'data/retiree_blacklist.csv'
    check_path(blacklist_path)
    industry = True # 產學|研究計畫(過濾本次申請人)
    apply_data = {}
    # 讀取資料
    apply_data = load_apply_data(apply_path, ['計畫書資料'])
    if industry:
        all_applicants_set = set(info['manager'] for info in apply_data.values() if info['manager'])
    else:
        all_applicants_set = set()
    pages = ['title', 'keywords', 'application_directions', 'problems_to_solve', 'goals_to_achieve', 'methods_to_solve']
    # pages = ['title']
    result = calculate_score(group_by_project_dict(load_score_data(input_score_file, pages), pages))
    # 合併資訊
    for proj, info in result.items():
        app = apply_data.get(proj, {'no':'', 'managerSchool': '', 'coManger': [], 'coSchool': []})
        info.update({'no':app['no'],'school': app['managerSchool'], 'co_managers': app['coManger'], 'co_schools': app['coSchool']})

    # 載入委員與黑名單
    reviewer_school_map = {}
    reviewer_advisor_map = {}
    if os.path.exists(committee_uni_file):
        for _, r in pd.read_excel(committee_uni_file, dtype=str).fillna("").iterrows():
            reviewer_school_map[str(r['名稱']).strip()] = str(r['現任學校']).strip()
            reviewer_advisor_map[str(r['名稱']).strip()] = str(r['博士指導教授']).strip().split('/')
    blacklist = pd.read_csv(blacklist_path)['姓名'].astype(str).tolist() if os.path.exists(blacklist_path) else []

    # 上色存檔
    save_results_with_highlight(result, 'data/research_proj/115計算機學門審查/瑞典/recommendation_results_org_colored(計畫書資料).xlsx', reviewer_school_map, reviewer_advisor_map, blacklist, all_applicants_set)
    # 過濾與紀錄理由
    for proj, info in result.items():
        filtered, reasons = [], []
        apply_advisors = reviewer_advisor_map.get(info['manager'], [])
        for cand in info['candidates']:
            if len(filtered) >= 10:
                break
            cand_name = cand['name']
            cand_school = reviewer_school_map.get(cand_name, "")
            cand['school'] = cand_school
            
            # 調用模組化判定
            reason_type, _ = check_conflict_status(cand_name, cand_school, apply_advisors, info, all_applicants_set, blacklist)
            if reason_type:
                reasons.append(f"{cand_name}:{reason_type}")
            else:
                filtered.append(cand)
        print(reasons)
        info['final_candidates'] = filtered
        if reasons:
            info['filter_reason_str'] = "[" + ";".join(reasons) + ";]"
        else:
            info['filter_reason_str'] = "[]"
    save_results_clean(result, 'data/research_proj/115計算機學門審查/瑞典/recommendation_results_filter(計畫書資料).xlsx')

if __name__ == "__main__":
    main()
