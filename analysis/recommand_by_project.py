import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from save_commitee_data import filiter_school 

# --- 1. 利益迴避核心判定模組 ---
def check_conflict_status(cand_name, cand_school, apply_advisors, app_info, all_applicants_set, blacklist):
    app_name = app_info.get('manager', '')
    app_school = app_info.get('school', '')
    co_managers = app_info.get('co_managers', [])
    co_schools = app_info.get('co_schools', [])

    if cand_name in blacklist: return "黑名單", "gray"
    if cand_name == app_name: return "本人", "green"
    if cand_name in all_applicants_set: return "為本次計畫申請者", "green"
    if cand_name in co_managers: return "是推薦人", "blue"
    if app_school and cand_school and app_school == cand_school: return "與申請人同校", "red"
    if cand_school and cand_school in co_schools: return "與推薦人同校", "blue"
    if cand_name in apply_advisors: return "為申請者博士指導教授", "red"

    return None, None

# --- 2. 讀取與計算邏輯 ---
def load_apply_data(file_path, pages):
    apply_dicts = {} 
    for page in pages:
        try:
            df = pd.read_excel(file_path, sheet_name=page, dtype=str).fillna("")
            for _, row in df.iterrows():
                title = str(row.get('研習主題') or row.get('計畫名稱') or "").strip()
                no =  str(row.get('計畫編號')).strip()
                if not title: continue
                school, dept = filiter_school(row.get('申請人機關名稱'))
                
                co_m, co_s = [], []
                co_manager_raw = row.get('共同主持人') 
                
                if pd.notna(co_manager_raw) and str(co_manager_raw).strip():
                    for item in str(co_manager_raw).split(';'):
                        match = re.match(r'^(.*?)[(（](.*)[)）]', item.strip())
                        if match:
                            name, s_raw = match.group(1).strip(), match.group(2).strip()
                            s_clean, _ = filiter_school(s_raw)
                            co_m.append(name); co_s.append(s_clean)
                        else:
                            co_m.append(item.strip())
                apply_dicts[title] = {
                    'no': no,
                    'manager': '',
                    'managerSchool': school,
                    'coManger': co_m,
                    'coSchool': co_s
                }
        except Exception as e:
            print(f"跳過頁面 {page}: {e}")
            continue
    return apply_dicts

def load_score_data(file_path, pages):
    page_data = {}
    for page in pages:
        try:
            df = pd.read_excel(file_path, sheet_name=page, dtype=str).fillna("")
            page_data[page] = df[['apply_project', 'manager', 'similarity_score', 'recommended_manager', 'recommended_project']]
        except Exception as e:
            print(f"跳過頁面 {page}: {e}")
            continue
    return page_data

def calculate_score(data):
    weights = {'title': 1, 'keyword': 1, 'application_directions': 3, 'problems_to_solve': 3, 'goals_to_achieve': 1, 'methods_to_solve': 1}
    total_w = sum(weights.values())
    results = {}
    
    for proj, pages_data in data.items():
        m_scores = {}
        app_m = ""
        
        for pg, rows in pages_data.items():
            w = weights.get(pg, 0)
            for r in rows:
                if not app_m: app_m = r.get('manager', '')
                rec = r.get('recommended_manager')
                rec_proj = r.get('recommended_project')
                
                if rec and rec_proj:
                    try: s = float(r.get('similarity_score', 0))
                    except ValueError: s = 0
                    
                    key = (rec, rec_proj)
                    m_scores[key] = m_scores.get(key, 0) + (s * w)
        
        best_scores = {}
        for (manager, project), score in m_scores.items():
            final_score = score / total_w
            if manager not in best_scores or final_score > best_scores[manager]['score']:
                best_scores[manager] = {'score': final_score, 'project': project}
        
        cand_list = sorted(
            [{'name': m, 'score': info['score'], 'recommended_project': info['project']} for m, info in best_scores.items()], 
            key=lambda x: x['score'], reverse=True
        )
        
        results[proj] = {'project_name': proj, 'manager': app_m, 'school': "", 'candidates': cand_list}
        
    return results

# --- 3. 存檔並上色 (更新欄位與上色邏輯) ---
def save_results_with_highlight(results, output_filename, reviewer_school_map, reviewer_advisor_map, blacklist, all_applicants_set):
    excel_rows = []
    
    for project, info in results.items():
        row = {"計畫編號": info['no'], "計畫名稱": project, '主持人': info['manager'], "申請機構": info['school']}
        
        # 預先初始化 10 個名額的欄位，確保 Excel 欄位順序不會因為人數不足而跑版
        for i in range(10):
            row[f"推薦委員 {i+1}"] = ""
            row[f"相似度分數 {i+1}"] = ""
            row[f"推薦委員學校 {i+1}"] = ""
            row[f"推薦專案 {i+1}"] = ""
            
        reasons_list = [] 
        apply_advisors = reviewer_advisor_map.get(info['manager'], [])
        
        for i, cand in enumerate(info.get('candidates', [])[:10]):
            cand_name = cand['name']
            cand_school = reviewer_school_map.get(cand_name, "")
            
            row[f"推薦委員 {i+1}"] = cand_name
            row[f"相似度分數 {i+1}"] = cand['score']
            row[f"推薦委員學校 {i+1}"] = cand_school
            row[f"推薦專案 {i+1}"] = cand.get('recommended_project', '') # 新增推薦專案
            
            reason_type, _ = check_conflict_status(cand_name, cand_school, apply_advisors, info, all_applicants_set, blacklist)
            if reason_type:
                reasons_list.append(f"{cand_name}:{reason_type}")
        
        row["過濾原因"] = "[" + ";".join(reasons_list) + ";]" if reasons_list else "[]"
        excel_rows.append(row)

    pd.DataFrame(excel_rows).to_excel(output_filename, index=False)

    # 上色處理
    wb = load_workbook(output_filename)
    ws = wb.active
    color_map = {
        "gray": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "green": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "blue": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "red": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    }

    for i, project in enumerate(results):
        row_idx = i + 2
        info = results[project]
        apply_advisors = reviewer_advisor_map.get(info['manager'], [])
        
        for k, cand in enumerate(info.get('candidates', [])[:10]):
            cand_name = cand['name']
            cand_school = reviewer_school_map.get(cand_name, "")
            _, color_key = check_conflict_status(cand_name, cand_school, apply_advisors, info, all_applicants_set, blacklist)
            
            if color_key:
                # 每個委員佔 4 個欄位 (名字, 分數, 學校, 專案)，起始欄位為 5
                base_col = 5 + (k * 4)
                for c in range(4): # 連同「推薦專案」一起上色
                    ws.cell(row=row_idx, column=base_col+c).fill = color_map[color_key]
    
    wb.save(output_filename)

# --- 4. 存檔乾淨名單 (更新欄位) ---
def save_results_clean(results, output_filename):
    excel_rows = []
    for project, info in results.items():
        row = {"計畫編號": info['no'], "計畫名稱": project, "主持人": info['manager'], "申請機構": info['school'], "過濾原因": info.get('filter_reason_str', "")}
        
        # 同樣預先初始化防跑版
        for i in range(10):
            row[f"推薦委員 {i+1}"] = ""
            row[f"相似度分數 {i+1}"] = ""
            row[f"推薦委員學校 {i+1}"] = ""
            row[f"推薦專案 {i+1}"] = ""
            
        for i, cand in enumerate(info.get('final_candidates', [])[:10]):
            row[f"推薦委員 {i+1}"] = cand['name']
            row[f"相似度分數 {i+1}"] = cand['score']
            row[f"推薦委員學校 {i+1}"] = cand.get('school', '') 
            row[f"推薦專案 {i+1}"] = cand.get('recommended_project', '') # 新增推薦專案
            
        excel_rows.append(row)
    pd.DataFrame(excel_rows).to_excel(output_filename, index=False)

def group_by_project_dict(datas, pages):
    results = {}
    for page in pages:
        if page not in datas: continue
        for row in datas[page].to_dict('records'):
            p = row.get('apply_project')
            if p:
                if p not in results: results[p] = {}
                if page not in results[p]: results[p][page] = []
                results[p][page].append(row)
    return results

def check_path(path):
    if not os.path.exists(path):
        print(f"找不到檔案: {path}")

# --- 5. 主程式 ---
def main():
    apply_path = 'data/research_proj/115計算機學門審查/新興大專生計畫/115 新興 大專生計畫-智慧計算.xlsx'
    check_path(apply_path)
    input_score_file = 'data/research_proj/115計算機學門審查/新興大專生計畫/result_score(115 新興 大專生計畫-智慧計算_by_project_research).xlsx'
    check_path(input_score_file)
    committee_uni_file = 'data/RDF_database/committee_all_education_with_advisor.xlsx'
    check_path(committee_uni_file)
    blacklist_path = 'data/retiree_blacklist.csv'
    check_path(blacklist_path)
    
    industry = True 
    apply_data = load_apply_data(apply_path, ['工作表1'])
    all_applicants_set = set(info['manager'] for info in apply_data.values() if info['manager']) if industry else set()
    
    # pages = ['title', 'keywords', 'application_directions', 'problems_to_solve', 'goals_to_achieve', 'methods_to_solve']
    pages = ['title']
    result = calculate_score(group_by_project_dict(load_score_data(input_score_file, pages), pages))
    for proj, info in result.items():
        app = apply_data.get(proj, {'no':'', 'managerSchool': '', 'coManger': [], 'coSchool': []})
        info.update({'no':app['no'],'school': app['managerSchool'], 'co_managers': app['coManger'], 'co_schools': app['coSchool']})

    reviewer_school_map = {}
    reviewer_advisor_map = {}
    if os.path.exists(committee_uni_file):
        for _, r in pd.read_excel(committee_uni_file, dtype=str).fillna("").iterrows():
            name = str(r['名稱']).strip()
            reviewer_school_map[name] = str(r['現任學校']).strip()
            
            # 優化：安全處理指導教授欄位，避免 nan 變字串
            advisor_str = str(r['博士指導教授']).strip()
            if advisor_str and advisor_str.lower() != 'nan':
                reviewer_advisor_map[name] = [adv.strip() for adv in advisor_str.split('/')]
            else:
                reviewer_advisor_map[name] = []
                
    blacklist = pd.read_csv(blacklist_path)['姓名'].astype(str).tolist() if os.path.exists(blacklist_path) else []

    save_results_with_highlight(result, 'data/research_proj/115計算機學門審查/新興大專生計畫/recommendation_results_org_colored(115 新興 大專生計畫-智慧計算_by_project_research).xlsx', reviewer_school_map, reviewer_advisor_map, blacklist, all_applicants_set)
    
    for proj, info in result.items():
        filtered, reasons = [], []
        apply_advisors = reviewer_advisor_map.get(info['manager'], [])
        for cand in info['candidates']:
            if len(filtered) >= 10: break
            cand_name = cand['name']
            cand_school = reviewer_school_map.get(cand_name, "")
            cand['school'] = cand_school
            
            reason_type, _ = check_conflict_status(cand_name, cand_school, apply_advisors, info, all_applicants_set, blacklist)
            if reason_type:
                reasons.append(f"{cand_name}:{reason_type}")
            else:
                filtered.append(cand)
                
        info['final_candidates'] = filtered
        info['filter_reason_str'] = "[" + ";".join(reasons) + ";]" if reasons else "[]"
        
    save_results_clean(result, 'data/research_proj/115計算機學門審查/新興大專生計畫/recommendation_results_filter(115 新興 大專生計畫-智慧計算_by_project_research).xlsx')

if __name__ == "__main__":
    main()
