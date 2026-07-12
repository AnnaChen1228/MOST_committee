from itertools import chain
import json
import shutil
import tqdm
import chromadb
import re
import pandas as pd
import numpy as np
import ast
import sys
import os
os.environ["TOKENIZERS_PARALLELISM"] = "false"

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from utils.cal_embedding_bge_zh import calculate_docs_embedding_zh, get_embeddings_zh
from utils.load_source_excel import get_project_df, get_industry_coop_proj
from langchain_community.vectorstores.chroma import Chroma
from utils.load_former_manager import get_former_manager

from utils.get_setting import setting_data, print_setting_data, find_key_path, value_of_key, get_run_output_dir
from utils.filter_method import *
from utils.store_vectordb import store_basic_vector_db, store_abstract_vector_db
from utils.generate_abstract import local_generate
class MissingFieldsException(Exception):
    pass

# 中間過程 JSON 存於 temp/ 子資料夾，最終輸出資料夾只留三支結果檔
_TEMP_SUBDIR = "temp"
REC_MAP_FILENAME = "推薦依據計畫對應.json"
REC_POOL_FILENAME = "推薦候選池.json"
FILTERED_TOP_FILENAME = "過濾後合格委員.json"

def _temp_path(run_dir, filename):
    path = os.path.join(run_dir, _TEMP_SUBDIR, filename)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    return path

def _make_rec_map_key(sheet_name, project_name, manager):
    '''建立 (分頁, 申請計畫名稱, 推薦委員) 的對應鍵值。'''
    return f"{sheet_name}|||{str(project_name).strip()}|||{str(manager).strip()}"

def _make_pool_key(sheet_name, project_name):
    '''建立 (分頁, 申請計畫名稱) 的對應鍵值。'''
    return f"{sheet_name}|||{str(project_name).strip()}"

def load_into_chroma_bge_manager(is_industry=False, progress_callback=None):
    # 讀取黑名單
    blacklist_csv_path = find_key_path("退休或黑名單委員")
    blacklist_df = pd.read_csv(blacklist_csv_path, encoding='utf-8')
    blacklist = blacklist_df['姓名'].tolist()
    print(f"黑名單或是已退休委員:", blacklist)
    
    # 決定資料庫路徑
    if is_industry:
        chroma_basic_path = find_key_path('CHROMA_BASIC_INDUSTRY')
        chroma_abstract_path = find_key_path('CHROMA_ABSTRACT_INDUSTRY')
        db_type = 'industry'
    else:
        chroma_basic_path = find_key_path('CHROMA_BASIC')
        chroma_abstract_path = find_key_path('CHROMA_ABSTRACT')
        db_type = 'research'

    # ================= 新增：清空舊的 Collection =================
    print("正在清理舊的向量資料庫...")
    if os.path.exists(chroma_basic_path):
        try:
            shutil.rmtree(chroma_basic_path)
            print(f"已成功刪除並重置資料夾: {chroma_basic_path}")
        except Exception as e:
            print(f"刪除 {chroma_basic_path} 失敗: {e}")

    # 清理 Abstract 資料庫
    if os.path.exists(chroma_abstract_path):
        try:
            shutil.rmtree(chroma_abstract_path)
            print(f"已成功刪除並重置資料夾: {chroma_abstract_path}")
        except Exception as e:
            print(f"刪除 {chroma_abstract_path} 失敗: {e}")
    # ============================================================

    # 取得對應的 Dataframes
    df_list = get_industry_coop_proj() if is_industry else get_project_df()

    project_data = {} # 用來存放給 store_basic_vector_db 和 store_abstract_vector_db 的資料
    manager_group = {} # 保留原本產出 JSON 的邏輯 (依主持人分組)

    for key, year_data in df_list.items():
        # key 通常是年度 (pass_year)
        # for i in tqdm.tqdm(range(len(year_data)), desc=str(key)):
        total_items = len(year_data)
        for i in tqdm.tqdm(range(len(year_data)), desc=key):
            # 【關鍵加入】如果 GUI 有傳入進度更新函式，就呼叫它來更新畫面
            if progress_callback:
                # 傳入：目前第幾筆(i+1), 總共幾筆, 目前的年份(key)
                progress_callback(i + 1, total_items, str(key))
            row = year_data.iloc[i]
            
            manager = str(row.get('計畫主持人', '')).strip()
            if not manager or manager.lower() == 'nan':
                continue
                
            project_name = str(row.get('計畫中文名稱', '') or '').strip()
            abstract = str(row.get('中文摘要', '') or '').strip()
            keywords = str(row.get('中文關鍵字', '') or '').strip()
            approved = str(row.get('通過', 'false')).strip().lower()
            
            # 產學資料以外，跳過未通過的計畫
            if not is_industry and approved != 'true':  
                continue
            
            # 跳過黑名單
            if manager in blacklist:
                continue
            
            # 呼叫 LLM 進行摘要拆解
            parsed_abstract = {}
            if abstract and abstract.lower() != 'nan':
                try:
                    # 呼叫您寫好的 local_generate
                    response = local_generate(abstract)
                    # 將 LLM 回傳的 JSON 字串轉為 Python 字典
                    parsed_abstract = json.loads(response.message.content)
                except json.JSONDecodeError:
                    print(f"警告: LLM 回傳的格式非有效 JSON，計畫名稱: {project_name}")
                except RuntimeError as e:
                    # 【關鍵修改】如果捕捉到伺服器超時的 RuntimeError，直接往上拋出！
                    # 這樣就會立刻中斷迴圈，並一路傳遞到 execute_mode 顯示錯誤視窗
                    raise e
                except Exception as e:
                    print(f"解析摘要發生錯誤，計畫名稱: {project_name}, 錯誤: {e}")

            # 準備寫入向量資料庫的資料格式 (以 project_name 為 key)
            if project_name:
                project_data[project_name] = {
                    'host_name': manager,
                    'pass_year': key,
                    'title': project_name,
                    'keywords': keywords,
                    'application_directions': parsed_abstract.get('application_directions', ''),
                    'problems_to_solve': parsed_abstract.get('problems_to_solve', ''),
                    'goals_to_achieve': parsed_abstract.get('goals_to_achieve', ''),
                    'methods_to_solve': parsed_abstract.get('methods_to_solve', '')
                }
            
            # 保留原本的 Manager Group 邏輯，用於輸出傳統的 JSON 檔案
            text_for_json = f"{project_name} {abstract} {keywords}\n"
            if manager in manager_group:
                manager_group[manager] += f"\n{text_for_json}"
            else:
                manager_group[manager] = text_for_json

    # 呼叫您自定義的儲存函式 (LangChain Chroma Wrapper)
    print(f"\n開始將 Basic 資料寫入 ChromaDB ({db_type})...")
    if progress_callback:
        # 傳入 0, 0，並寫上提示文字
        progress_callback(0, 0, f"正在將 Basic 資料寫入資料庫...")
        
    store_basic_vector_db(db_type, project_data)

    print(f"開始將 Abstract 資料寫入 ChromaDB ({db_type})...")
    if progress_callback:
        progress_callback(0, 0, f"正在將 Abstract 資料寫入資料庫...")
        
    store_abstract_vector_db(db_type, project_data)

    # 決定 JSON 儲存路徑並儲存 (保留原本的 BGE_MANAGER JSON 輸出)
    bge_manager_key = "BGE_INDUSTRY_MANAGER" if is_industry else "BGE_MANAGER"
    bge_manager_path = find_key_path(bge_manager_key)
        
    with open(bge_manager_path, 'w', encoding='utf-8') as f:
        json.dump(manager_group, f, ensure_ascii=False)
    print(f"已成功儲存主持人 JSON 至: {bge_manager_path}")

    pass_data_store_key = "PASS_INDUSTRY" if is_industry else "PASS_RESEARCH"
    pass_data_store_path = find_key_path(pass_data_store_key)
    # 1. 先將所有資料轉成 DataFrame
    df = pd.DataFrame(list(project_data.values()))

    # 2. 使用 ExcelWriter 來建立多個 Sheet
    # engine='openpyxl' 是寫入 .xlsx 必備的引擎
    with pd.ExcelWriter(pass_data_store_path, engine='openpyxl') as writer:
        
        # 3. 依照 'pass_year' 進行分組 (groupby)
        for year, group_df in df.groupby('pass_year'):
            
            # 確保 Sheet 名稱是字串 (例如 "112", "113")
            sheet_name = str(year)
            
            # 將該年份的資料寫入對應的 Sheet 中
            group_df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"已成功儲存通過案件資料至: {pass_data_store_path}")

def search_v3(is_industry=False, progress_callback=None):

    tabs = value_of_key("計畫SHEET")
    output_excel_folder_path = find_key_path("統計表分析")

    # 研究計畫 DB 為預設；產學模式額外加入產學 DB（各輸出獨立 sheet）
    excel_folder_path = find_key_path("產學合作申請名冊") if is_industry else find_key_path("研究計畫申請名冊")
    search_configs = [
        {
            'label':    '研究計畫' if is_industry else None,
            'basic':    find_key_path('CHROMA_BASIC'),
            'abstract': find_key_path('CHROMA_ABSTRACT'),
        }
    ]
    if is_industry:
        search_configs.append({
            'label':    '產學合作',
            'basic':    find_key_path('CHROMA_BASIC_INDUSTRY'),
            'abstract': find_key_path('CHROMA_ABSTRACT_INDUSTRY'),
        })

    former_manager = get_former_manager(find_key_path("曾任委員"))

    WEIGHTS = {
        'title': 1, 'keywords': 1,
        'application_directions': 3, 'problems_to_solve': 3,
        'goals_to_achieve': 1, 'methods_to_solve': 1,
    }
    TOTAL_W = sum(WEIGHTS.values())

    other_fields = value_of_key("計畫相關其他欄位")
    required_fields = []
    for key in ["計畫名稱", "中文關鍵字", "計劃摘要", "職稱", "申請主持人欄位名稱", "申請機構欄位名稱"]:
        temp_value = value_of_key(key)
        if temp_value is not None and temp_value != '':
            required_fields.append(temp_value)
    for key in ["申請共同主持人"]:
        for v in value_of_key(key):
            if v not in required_fields and v != '':
                required_fields.append(v)
    filter_fields = required_fields + [f for f in other_fields if f not in required_fields]

    RECOMMAND_AMOUNT = 10   # 原始表顯示的推薦委員數量（欄位數，請勿隨意更動，著色邏輯有相依）
    POOL_SIZE        = 30   # 候選池大小：過濾後若不足 RECOMMAND_AMOUNT 位會從此池遞補
    SELECT_AMOUNT    = 3
    SELECT_BOX_SYMBOL = ['Y', 'Z', 'AA']

    xls    = pd.ExcelFile(excel_folder_path)
    writer = pd.ExcelWriter(output_excel_folder_path, engine='openpyxl')

    run_output_dir = get_run_output_dir() or "./data/output"
    similarity_record_path = os.path.join(run_output_dir, f"similarity_record_{value_of_key('FINAL_COMMITTEE')}")
    os.makedirs(os.path.dirname(similarity_record_path), exist_ok=True)
    # 推薦委員「依據哪一個過去計畫被推薦」：key = (分頁, 申請計畫, 委員) -> 過去計畫名稱
    recommend_project_map = {}
    # 較深的候選池（供過濾後遞補使用）：key = (分頁, 申請計畫) -> [{manager, score, project}, ...]
    recommend_pool_map = {}
    similarity_rows = []
    parsed_rows  = []
    parse_cache  = {}  # { tab: [parsed_abstract_dict, ...] }

    project_name_field_name    = value_of_key("計畫名稱")
    chinese_keyword_field_name = value_of_key("中文關鍵字")
    abstract_field_name        = value_of_key("計劃摘要")
    apply_manager_field_name   = value_of_key("申請主持人欄位名稱")

    try:
        # ── Phase 1：LLM 拆解所有申請資料（全部跑完再進比對）──
        for tab in tabs:
            df_parse = pd.read_excel(xls, tab)
            parse_cache[tab] = []
            total_parse = len(df_parse)
            for i in tqdm.tqdm(range(total_parse), desc=f"摘要拆解({tab})"):
                project_name  = str(df_parse.iloc[i].get(project_name_field_name,   '') or '').strip()
                keywords      = str(df_parse.iloc[i].get(chinese_keyword_field_name, '') or '').strip()
                abstract      = str(df_parse.iloc[i].get(abstract_field_name,        '') or '').strip()
                apply_manager = str(df_parse.iloc[i].get(apply_manager_field_name,   '') or '').strip()
                parsed_abstract = {}
                if abstract and abstract.lower() != 'nan':
                    try:
                        response = local_generate(abstract)
                        parsed_abstract = json.loads(response.message.content)
                    except RuntimeError as e:
                        raise e
                    except Exception as e:
                        print(f"解析摘要失敗: {project_name}, 錯誤: {e}")
                parse_cache[tab].append(parsed_abstract)
                parsed_rows.append({
                    'tab':                    tab,
                    'apply_project':          project_name,
                    'apply_manager':          apply_manager,
                    'keywords':               keywords,
                    'abstract':               abstract,
                    'application_directions': parsed_abstract.get('application_directions', ''),
                    'problems_to_solve':      parsed_abstract.get('problems_to_solve', ''),
                    'goals_to_achieve':       parsed_abstract.get('goals_to_achieve', ''),
                    'methods_to_solve':       parsed_abstract.get('methods_to_solve', ''),
                })
                if progress_callback:
                    progress_callback(i + 1, total_parse, f"摘要拆解({tab})")

        pd.DataFrame(parsed_rows).to_excel(find_key_path("申請計畫拆解結果"), index=False)

        # ── Phase 2：向量相似度比對（每個 DB config 各跑一輪）──
        for cfg in search_configs:
            db_label = cfg['label']
            if progress_callback:
                progress_callback(0, 0, f"進行{db_label}資料庫的相似度比對...")
            embedding_fn = get_embeddings_zh()
            vectorstores = {}
            for col in ['title', 'keywords']:
                vectorstores[col] = Chroma(
                    collection_name=col,
                    persist_directory=cfg['basic'],
                    embedding_function=embedding_fn
                )
            for col in ['application_directions', 'problems_to_solve', 'goals_to_achieve', 'methods_to_solve']:
                vectorstores[col] = Chroma(
                    collection_name=col,
                    persist_directory=cfg['abstract'],
                    embedding_function=embedding_fn
                )

            for tab in tabs:
                sheet_name = f"{tab}_{db_label}" if db_label else tab
                page_manager_list = []

                df = pd.read_excel(xls, tab)

                existing_fields = df.columns.tolist()
                missing_fields  = [f for f in filter_fields if f not in existing_fields]
                if missing_fields:
                    print("現有欄位:", existing_fields)
                    print("應當欄位:", filter_fields)
                    raise ValueError("欄位不匹配，程式碼運行停止")

                df = df[filter_fields]

                for i in range(RECOMMAND_AMOUNT):
                    df['推薦委員' + str(i + 1)] = ''
                    df['相關分數' + str(i + 1)] = ''
                df['前任委員占比'] = ''
                for i in range(SELECT_AMOUNT):
                    df['選取委員' + str(i + 1)] = ''

                total_items = len(df)
                for i in tqdm.tqdm(range(total_items), desc=f"{sheet_name}"):
                    manager_list    = []
                    project_name    = str(df.iloc[i].get(project_name_field_name,   '') or '').strip()
                    keywords        = str(df.iloc[i].get(chinese_keyword_field_name, '') or '').strip()
                    apply_manager   = str(df.iloc[i].get(apply_manager_field_name,   '') or '').strip()
                    parsed_abstract = parse_cache[tab][i]

                    query_texts = {
                        'title':                  project_name,
                        'keywords':               keywords,
                        'application_directions': parsed_abstract.get('application_directions', ''),
                        'problems_to_solve':      parsed_abstract.get('problems_to_solve', ''),
                        'goals_to_achieve':       parsed_abstract.get('goals_to_achieve', ''),
                        'methods_to_solve':       parsed_abstract.get('methods_to_solve', ''),
                    }

                    manager_proj_scores = {}
                    for field, query_text in query_texts.items():
                        if not query_text or not query_text.strip():
                            continue
                        w = WEIGHTS[field]
                        results = vectorstores[field].similarity_search_with_relevance_scores(
                            query_text, k=RECOMMAND_AMOUNT * 3
                        )
                        for doc, score in results:
                            key = (doc.metadata['manager'], doc.metadata['title'])
                            manager_proj_scores[key] = manager_proj_scores.get(key, 0) + score * w
                            similarity_rows.append({
                                'db_source':           db_label or '研究計畫',
                                'apply_project':       project_name,
                                'apply_manager':       apply_manager,
                                'field':               field,
                                'field_weight':        w,
                                'query_text':          query_text,
                                'recommended_manager': doc.metadata['manager'],
                                'recommended_project': doc.metadata['title'],
                                'compared_text':       doc.page_content,
                                'field_score':         score,
                            })

                    best_per_manager = {}
                    for (rec_manager, rec_proj), weighted_sum in manager_proj_scores.items():
                        final_score = weighted_sum / TOTAL_W
                        if rec_manager not in best_per_manager or final_score > best_per_manager[rec_manager]['score']:
                            best_per_manager[rec_manager] = {'score': final_score, 'project': rec_proj}

                    ranked_managers = sorted(
                        best_per_manager.items(),
                        key=lambda x: x[1]['score'],
                        reverse=True
                    )
                    # 較深的候選池（供過濾後遞補），原始表仍只顯示前 RECOMMAND_AMOUNT 名
                    pool_managers   = ranked_managers[:POOL_SIZE]
                    sorted_managers = ranked_managers[:RECOMMAND_AMOUNT]

                    # 整池都記錄「推薦依據計畫」，遞補進來的委員之後也能顯示依據
                    pool_records = []
                    for rec_manager, info in pool_managers:
                        recommend_project_map[_make_rec_map_key(sheet_name, project_name, rec_manager)] = info['project']
                        pool_records.append({
                            'manager': rec_manager,
                            'score':   info['score'],
                            'project': info['project'],
                        })
                    recommend_pool_map[_make_pool_key(sheet_name, project_name)] = pool_records

                    for j, (rec_manager, info) in enumerate(sorted_managers):
                        df.loc[df.index[i], '推薦委員' + str(j + 1)] = rec_manager
                        df.loc[df.index[i], '相關分數' + str(j + 1)] = info['score']
                        manager_list.append(rec_manager)

                    page_manager_list.append(manager_list)
                    df.loc[df.index[i], '前任委員占比'] = len([x for x in manager_list if x in former_manager]) / RECOMMAND_AMOUNT

                df.to_excel(writer, sheet_name=sheet_name, index=False)

                workbook  = writer.book
                worksheet = workbook[sheet_name]

                for j in range(SELECT_AMOUNT):
                    for i, manager_list in enumerate(page_manager_list):
                        data_range = ','.join(manager_list)
                        dv = DataValidation(type="list", formula1=f'"{data_range}"', allow_blank=True)
                        dv.add(SELECT_BOX_SYMBOL[j] + str(i + 2))
                        worksheet.add_data_validation(dv)

                highligh_former_manager(writer, sheet_name, former_manager, output_excel_folder_path)
                draw_color_for_similarity_score(writer, sheet_name, output_excel_folder_path)

    except Exception as e:
        if not writer.book.sheetnames:
            print("ERROR")
            writer.book.create_sheet(title="Error")
        raise

    finally:
        writer.close()
        # 輸出「推薦依據計畫對應表」供後續加上註解使用
        rec_map_path = _temp_path(run_output_dir, REC_MAP_FILENAME)
        with open(rec_map_path, 'w', encoding='utf-8') as f:
            json.dump(recommend_project_map, f, ensure_ascii=False)
        # 輸出「候選池」供 filter_committee 做過濾與遞補
        rec_pool_path = _temp_path(run_output_dir, REC_POOL_FILENAME)
        with open(rec_pool_path, 'w', encoding='utf-8') as f:
            json.dump(recommend_pool_map, f, ensure_ascii=False)
        similarity_df = pd.DataFrame(similarity_rows)
        fields = ['title', 'keywords', 'application_directions', 'problems_to_solve', 'goals_to_achieve', 'methods_to_solve']
        with pd.ExcelWriter(similarity_record_path, engine='openpyxl') as similarity_writer:
            for field in fields:
                field_df = similarity_df[similarity_df['field'] == field] if not similarity_df.empty else pd.DataFrame()
                field_df.to_excel(similarity_writer, sheet_name=field, index=False)

def draw_color_for_similarity_score(writer, tab, output_excel):
    
    from openpyxl.formatting.rule import ColorScaleRule
    
    SIMILARITY_SCORE_RANGE = '$E$2:$w$1000'
    workbook = writer.book
    worksheet = workbook[tab]
    rule = ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="F9F900")
    worksheet.conditional_formatting.add(SIMILARITY_SCORE_RANGE, rule)
    workbook.save(output_excel)

def highligh_former_manager(writer, tab, former_manager, output_excel):
    
    from openpyxl.formatting import Rule
    from openpyxl.styles.differential import DifferentialStyle
    
    RECOMMAND_MANAGER_SYMBOL = ['D','F','H','J','L','N','P','R','T','V']
    workbook = writer.book
    worksheet = workbook[tab]
    redFill = PatternFill(start_color='FFA042', end_color='FFA042', fill_type='solid')

    for s in RECOMMAND_MANAGER_SYMBOL:
        col1 = worksheet[s]
        for i, cell in enumerate(col1):
            cell_value = cell.value
            if cell_value in former_manager:
                rule = Rule(type="cellIs", operator="equal", formula=[f'"{cell_value}"'], dxf=DifferentialStyle(fill=redFill))
                # rule = Rule(type="expression", operator="equal", formula=[f'"{cell_value}"'], dxf=DifferentialStyle(fill=redFill))
                worksheet.conditional_formatting.add(f'{s}{str(i+1)}', rule)

    workbook.save(output_excel)

def statistic_committee():
    
    apply_project_file_year = value_of_key("計畫過去申請案件年分範圍")

    # 統計清單（比較清單）為選用：未提供或檔案不存在時，跳過「研究計劃（統計案件）」這個來源
    statistic_value = value_of_key("統計清單")
    statistic_excel_file = None
    statistic_sheets = []
    if statistic_value:
        statistic_folder_path = find_key_path("統計清單")
        if os.path.exists(statistic_folder_path):
            statistic_excel_file = pd.ExcelFile(statistic_folder_path)
            statistic_sheets = statistic_excel_file.sheet_names
        else:
            print(f"警告: 找不到統計清單檔案 {statistic_folder_path}，將跳過研究計劃（統計案件）來源。")
    else:
        print("提示: 未提供統計清單（比較清單），將跳過研究計劃（統計案件）來源。")

    industry_folder_path = find_key_path("產學過去申請名冊")
    industry_data = pd.read_excel(industry_folder_path)
    
    past_apply_project_path = find_key_path("計畫過去申請案件")
    past_apply_project_file = pd.ExcelFile(past_apply_project_path)
    
    newest_person_path = find_key_path("暫存最新人才資料庫")
    newest_person_df = pd.read_excel(newest_person_path)
    
    # committee_study_folder_path = find_key_path("碩博士論文_RDF")
    # committee_study_data = pd.read_excel(committee_study_folder_path)
    
    #@ 處理委員的所有相關學校名單: 名稱 - 年份 - 學校 - 職稱
    committee_person_RDF = []
    
    # #: 碩博士論文
    # for index, row in tqdm.tqdm(committee_study_data.iterrows(), desc="碩博士論文"):
    #     # 先檢查欄位是否存在
    #     name = row.get('學生姓名', '')
    #     year = row.get('畢業學年度', '') if '畢業學年度' in row else ''
    #     institution = row.get('畢業學校', '') if '畢業學校' in row else ''

    #     # 確保數據不為 NaN（轉為字串或預設值）
    #     year = str(year) if pd.notna(year) else ''
    #     institution = str(institution) if pd.notna(institution) else ''

    #     committee_person_RDF.append({
    #         '名稱': name,
    #         '年份': year,
    #         '機關名稱': institution,
    #         '職稱': ""
    #     })
    
    #: 暫存最新人才資料庫
    for index, row in newest_person_df.iterrows():
        committee_person_RDF.append({
            '名稱': row['名稱'],
            '年份': int(row['年份']),
            '機關名稱': row['機關名稱'],
            '職稱': row['職稱'],
            "來源": row['來源']
        })

    #: 研究計劃（申請案件）
    for year in apply_project_file_year:
        current_sheet = year
        past_apply_project_df = pd.read_excel(past_apply_project_file, current_sheet)
        for index, row in tqdm.tqdm(past_apply_project_df.iterrows(), desc=f"{current_sheet}"):
            committee_person_RDF.append({
                '名稱': row['計畫主持人'],
                '年份': year,
                '機關名稱': row['機關名稱'],
                '職稱': row.get('職稱', '教授'),
                "來源": f"研究計劃（申請案件）- {current_sheet}"
            })
        
        
    #: 研究計劃（統計案件）— 未提供統計清單則整段跳過
    if statistic_excel_file is not None:
        for year in apply_project_file_year:
            current_sheet = f"{year}總計畫清單"
            if current_sheet not in statistic_sheets:
                print(f"警告: 統計清單中找不到「{current_sheet}」分頁，跳過該年度統計案件來源。")
                continue
            statistic_df = pd.read_excel(statistic_excel_file, current_sheet)
            for index, row in tqdm.tqdm(statistic_df.iterrows(), desc=f"{current_sheet}"):
                committee_person_RDF.append({
                    '名稱': row['計畫主持人'],
                    '年份': year,
                    '機關名稱': row['機關名稱'],
                    '職稱': row.get('職稱', '教授'),
                    "來源": f"研究計劃（統計案件）- {current_sheet}"
                })
            
    #: 產學合作
    for index, row in industry_data.iterrows():
        committee_person_RDF.append({
            '名稱': row['計畫主持人'],
            '年份': row["申請年度"],
            '機關名稱': row['單位名稱'],
            '職稱': row['職稱'],
            # "來源": f"產學合作 - 序號:{row['序號']}"
        })
    
    committee_person_RDF_df = pd.DataFrame(committee_person_RDF)
    mask = (committee_person_RDF_df['機關名稱'].notna()) & (committee_person_RDF_df['機關名稱'] != '')
    committee_person_RDF_df = committee_person_RDF_df[mask]
    committee_person_RDF_df[['學校', '系所']] = committee_person_RDF_df['機關名稱'].apply(split_institution)
    committee_person_RDF_df = committee_person_RDF_df.sort_values(by=["名稱"])
    committee_person_RDF_df.to_excel(find_key_path("統計清單人才資料_RDF"), index=False)
    
    from utils.filter_method import extract_max_year
    
    committee_person_RDF_df['年份'] = committee_person_RDF_df['年份'].apply(extract_max_year)
    committee_person_RDF_df['年份'] = committee_person_RDF_df['年份'].fillna(0) 
    committee_person_RDF_df['年份'] = committee_person_RDF_df['年份'].astype(int)
    
    unique_person_RDF_df = committee_person_RDF_df.loc[committee_person_RDF_df.groupby('名稱')['年份'].idxmax()]
    unique_person_RDF_df.to_excel(find_key_path("統計清單人才資料_RDF_UNI"), index=False)

def filter_committee(is_industry=False):
    
    #: Load the data
    crawler_RDF_folder_path = find_key_path("碩博士論文_RDF")
    crawler_RDF_data = pd.read_excel(crawler_RDF_folder_path)

    statistical_analysis_folder_path = find_key_path("統計表分析")
    statistical_analysis_file = pd.ExcelFile(statistical_analysis_folder_path)

    if is_industry: apply_list_folder_path = find_key_path("產學合作申請名冊")
    else: apply_list_folder_path = find_key_path("研究計畫申請名冊")

    apply_list_file = pd.ExcelFile(apply_list_folder_path)

    committee_person_path = find_key_path("統計清單人才資料_RDF_UNI")
    committee_person_RDF_df = pd.read_excel(committee_person_path)

    #: 載入委員指導教授資料（路徑由 setting.yaml 的「委員指導教授資料」決定）
    advisor_data_path = find_key_path("委員指導教授資料")
    advisor_lookup = {}   # {人名 -> set(指導教授名稱)}
    student_lookup = {}   # {指導教授名稱 -> set(學生名稱)}
    if os.path.exists(advisor_data_path):
        advisor_df = pd.read_excel(advisor_data_path)
        for _, adv_row in advisor_df.iterrows():
            person = str(adv_row.get('名稱', '')).strip()
            if not person or person == 'nan':
                continue
            advisors = set()
            for col in ['博士指導教授', '碩士指導教授']:
                raw = str(adv_row.get(col, '')).replace('\xa0', '').strip()
                if not raw or raw in ('nan', '未找到'):
                    continue
                # 以 ／ 拆分多位指導教授（爬蟲格式：姓名１／姓名２）
                for part in raw.split('／'):
                    part = part.strip()
                    if len(part) >= 2:
                        advisors.add(part)
            if advisors:
                advisor_lookup[person] = advisor_lookup.get(person, set()) | advisors
        for student, advisors in advisor_lookup.items():
            for adv in advisors:
                student_lookup.setdefault(adv, set()).add(student)
    else:
        print(f"警告: 找不到委員指導教授資料檔案 {advisor_data_path}，將跳過師生關係過濾。")

    #: 載入 search_v3 產出的較深候選池（供過濾後遞補）
    run_output_dir = get_run_output_dir() or "./data/output"
    rec_pool_path = os.path.join(run_output_dir, REC_POOL_FILENAME)
    recommend_pool_map = {}
    if os.path.exists(rec_pool_path):
        with open(rec_pool_path, 'r', encoding='utf-8') as f:
            recommend_pool_map = json.load(f)
    BACKFILL_TARGET = 10  # 過濾後每案要補到的合格委員人數
    filtered_top_map = {}  # key = (分頁, 申請計畫) -> [{manager, score}, ...]（已遞補、最多 BACKFILL_TARGET 位）

    #- Strategy
    writer = pd.ExcelWriter(find_key_path("過濾相近後統計表"), engine='openpyxl')

    #@ 審查委員不能與計劃申請學校有關
    for stat_sheet in statistical_analysis_file.sheet_names:
        current_sheet_statistical_excel_data = pd.read_excel(statistical_analysis_file, sheet_name=stat_sheet)
        result_dict = []
        all_apply_members = current_sheet_statistical_excel_data[value_of_key("申請主持人欄位名稱")].to_list() # 所有申請人

        for index, statistical_row in current_sheet_statistical_excel_data.iterrows():
        # ~ 每個統計表的 row.
        
            # = 審查委員的背景（改用較深的候選池，過濾後才有足夠人選遞補）
            project_name_value = statistical_row[value_of_key("計畫名稱")]
            pool_records = recommend_pool_map.get(_make_pool_key(stat_sheet, project_name_value), [])
            # 候選人依分數排序（pool 本來就照分數），記錄順序與分數供遞補使用
            pool_order        = [rec['manager'] for rec in pool_records]
            pool_score_lookup = {rec['manager']: rec['score'] for rec in pool_records}

            committee_person_dict = []
            for rec in pool_records:
                committee_name = rec['manager']

                # 委員過去待過的學校
                been_list = []
                find_temp_df = committee_person_RDF_df[committee_person_RDF_df["名稱"] == committee_name]
                for index, row in find_temp_df.iterrows():
                    been_list.append(row["學校"])
                been_list = list(set(been_list))

                # 委員過去畢業的學校（以真正的委員姓名查詢）
                graduate_list = list(set(find_crawler_person_relative_school(committee_name, crawler_RDF_data)))

                # 委員職稱（人才資料庫_UNI 每人一列，取該列職稱）
                committee_title = find_temp_df['職稱'].iloc[0] if not find_temp_df.empty else ""

                committee_person_dict.append({
                    "委員名稱": committee_name,
                    "委員曾就職學校": been_list,
                    "委員過去畢業學校": graduate_list,
                    "職稱": committee_title
                })
                
            # = 申請學校 + 主持人學校 + 共同主持人學校
            '''
            如果有重疊，就會被歸類到 Filtered Members 裏
            反之則留在 Remaining Members
            至於 Filter Reasons 則紀錄每個被篩掉委員的具體原因
            '''
            total_committee_person_dict_result = {
                'Filtered Members': [],
                'Remaining Members': [],
                'Filter Reasons': {}
            }
            
            for sheet in apply_list_file.sheet_names: 
                current_sheet_apply_excel_data = pd.read_excel(apply_list_file, sheet_name=sheet)
                find_temp_df = current_sheet_apply_excel_data[
                    current_sheet_apply_excel_data[value_of_key("計畫名稱")] == statistical_row[value_of_key("計畫名稱")]
                ]
                
                for index, row in find_temp_df.iterrows(): 
                    joint_person_list = row.get(value_of_key("申請共同主持人"), pd.Series()).tolist()
                    joint_department_list = row.get(value_of_key("申請共同機構欄位名稱"), pd.Series()).tolist() 
                    common_person_dict = extract_text_in_parentheses(joint_person_list)
                    common_department_dict = extract_text_in_parentheses(joint_department_list)
                    
                    common_joint_list = common_person_dict + common_department_dict
                    
                    # 找到關聯性
                    project_manager_school = list([find_crawler_person_relative_school(name, crawler_RDF_data) for name, department in common_joint_list])
                    apply_school = {
                        "申請人名稱": row.get(value_of_key("申請主持人欄位名稱"), ''),
                        "申請人職稱": row.get(value_of_key("職稱"), ''),
                        "計畫申請學校": split_institution(row.get(value_of_key("申請機構欄位名稱"), ''))[0], 
                        "共同計畫主持的學校": [split_institution(department)[0] for name, department in common_joint_list],
                        "計畫主持人過去畢業的學校": find_crawler_person_relative_school(value_of_key("申請主持人欄位名稱"), crawler_RDF_data),
                        "共同主持人過去畢業的學校": list(chain.from_iterable(project_manager_school)), 
                    }
                    
                    #~ 審查委員不能與計劃申請學校(包含共同主持人)有關
                    # print("apply_school\n", apply_school)
                    # print("committee_person_dict\n", committee_person_dict)
                    filter_pairs = [("計畫申請學校", "委員曾就職學校"), ("共同計畫主持的學校", "委員曾就職學校")]
                    TITLE_RESTRICTIONS = {
                        "助理教授": ["教授", "研究員"], # 助理教授不能審教授或研究員
                        "助研究員": ["教授", "研究員"] # 助研究員不能審教授或研究員
                    }
                    current_committee_person_dict_result = filter_committee_advanced(
                        apply_school,
                        committee_person_dict,
                        filter_pairs,
                        all_apply_members,
                        TITLE_RESTRICTIONS,
                        whether_to_execute_the_option= {
                            "是否過濾申請人": True if is_industry else False,
                            "是否過濾相同學校": True,
                            "是否過濾職稱": True,
                            "是否過濾掉自身" : True
                        }
                    )
                    total_committee_person_dict_result = merge_committee_advanced(total_committee_person_dict_result, current_committee_person_dict_result)

                    #~ 師生關係過濾：申請人的指導教授 / 曾受申請人指導者 不得審查
                    applicant_name_str = str(row.get(value_of_key("申請主持人欄位名稱"), '')).strip()
                    advisor_filter_result = filter_committee_by_advisor(
                        applicant_name_str, committee_person_dict, advisor_lookup, student_lookup
                    )
                    total_committee_person_dict_result = merge_committee_advanced(total_committee_person_dict_result, advisor_filter_result)
                    
                    
                if len(find_temp_df) > 0: break #= 找不到東西，跳掉
        
            #- Input Selector
            # final_committee_person_list = [item for item in committee_person_dict["Remaining Members"][:3]]
            # for index, name in enumerate(final_committee_person_list):
            #     statistical_row[f"選取委員{index+1}"] = name
            
            #- Reason
            statistical_row["篩掉人員"] = total_committee_person_dict_result["Filtered Members"]
            statistical_row["篩選原因"] = total_committee_person_dict_result["Filter Reasons"]

            #- 過濾後遞補：依分數順序取未被篩掉者，補到 BACKFILL_TARGET 位
            filtered_out = set(total_committee_person_dict_result["Filtered Members"])
            backfilled = [
                {"manager": m, "score": pool_score_lookup.get(m, "")}
                for m in pool_order if m not in filtered_out
            ][:BACKFILL_TARGET]
            filtered_top_map[_make_pool_key(stat_sheet, project_name_value)] = backfilled

            result_dict.append(statistical_row)

        pd.DataFrame(result_dict).to_excel(writer, sheet_name=stat_sheet, index=False)
    writer.close()

    #: 輸出過濾後遞補的合格委員名單，供 excel_process_VBA 建立「_過濾後」分頁
    filtered_top_path = _temp_path(run_output_dir, FILTERED_TOP_FILENAME)
    with open(filtered_top_path, 'w', encoding='utf-8') as f:
        json.dump(filtered_top_map, f, ensure_ascii=False)

def filter_committee_remove(is_industry=False):
    #: 1. 預先載入資料 (移出迴圈以提升效能)
    crawler_RDF_folder_path = find_key_path("碩博士論文_RDF")
    crawler_RDF_data = pd.read_excel(crawler_RDF_folder_path)
    
    statistical_analysis_folder_path = find_key_path("統計表分析") 
    statistical_analysis_file = pd.ExcelFile(statistical_analysis_folder_path)
    
    # 根據是否為產學合作選擇路徑
    apply_list_folder_path = find_key_path("產學合作申請名冊") if is_industry else find_key_path("研究計畫申請名冊")
    apply_list_file = pd.ExcelFile(apply_list_folder_path)
    
    committee_person_path = find_key_path("統計清單人才資料_RDF_UNI")
    committee_person_RDF_df = pd.read_excel(committee_person_path)
    
    # 建立寫入器
    writer = pd.ExcelWriter(find_key_path("過濾相近後統計表"), engine='openpyxl')
    
    #@ 2. 開始處理每個分頁
    for sheet in statistical_analysis_file.sheet_names:
        current_sheet_statistical_excel_data = pd.read_excel(statistical_analysis_file, sheet_name=sheet)
        result_dict = []
        
        # 取得所有申請人清單
        all_apply_members = current_sheet_statistical_excel_data[value_of_key("申請主持人欄位名稱")].to_list()
        
        for index, statistical_row in current_sheet_statistical_excel_data.iterrows():
            # = 收集 30 位委員的背景資料
            committee_person_candidates = []
            for i in range(1, 31):  # 調整為 30 位
                col_name = f'推薦委員{i}'
                if col_name not in statistical_row or pd.isna(statistical_row[col_name]):
                    continue
                
                name = statistical_row[col_name]
                
                # 委員過去就職學校
                find_temp_df = committee_person_RDF_df[committee_person_RDF_df["名稱"] == name]
                been_list = list(set(find_temp_df["學校"].tolist()))
                
                # 委員畢業學校
                graduate_list = list(set(find_crawler_person_relative_school(name, crawler_RDF_data)))
                
                committee_person_candidates.append({
                    "委員名稱": name,
                    "委員曾就職學校": been_list,
                    "委員過去畢業學校": graduate_list,
                    "職稱": find_temp_df['職稱'].iloc[0] if not find_temp_df.empty else "未知"
                })
            
            # = 取得申請案相關資訊 (學校、主持人等)
            total_filter_result = {
                'Filtered Members': [],
                'Remaining Members': [],
                'Filter Reasons': {}
            }
            
            # 遍歷申請名冊找到對應計畫
            for app_sheet in apply_list_file.sheet_names:
                current_app_data = pd.read_excel(apply_list_file, sheet_name=app_sheet)
                target_project = current_app_data[current_app_data[value_of_key("計畫名稱")] == statistical_row[value_of_key("計畫名稱")]]
                
                if target_project.empty:
                    continue

                for _, app_row in target_project.iterrows():
                    # 提取共同主持人資訊
                    joint_person_list = app_row.get(value_of_key("申請共同主持人"), [])
                    joint_dept_list = app_row.get(value_of_key("申請共同機構欄位名稱"), [])
                    
                    common_joint_list = extract_text_in_parentheses(joint_person_list) + extract_text_in_parentheses(joint_dept_list)
                    
                    # 建立申請端學校清單
                    apply_school_info = {
                        "申請人名稱": app_row.get(value_of_key("申請主持人欄位名稱"), ''),
                        "計畫申請學校": split_institution(app_row.get(value_of_key("申請機構欄位名稱"), ''))[0],
                        "共同計畫主持的學校": [split_institution(dept)[0] for _, dept in common_joint_list],
                        "計畫主持人過去畢業的學校": find_crawler_person_relative_school(app_row.get(value_of_key("申請主持人欄位名稱")), crawler_RDF_data),
                    }

                    # 執行進階過濾
                    filter_pairs = [("計畫申請學校", "委員曾就職學校"), ("共同計畫主持的學校", "委員曾就職學校")]
                    TITLE_RESTRICTIONS = {"助理教授": ["教授", "研究員"], "助研究員": ["教授", "研究員"]}
                    
                    current_res = filter_committee_advanced(
                        apply_school_info, 
                        committee_person_candidates, 
                        filter_pairs, 
                        all_apply_members, 
                        TITLE_RESTRICTIONS,
                        whether_to_execute_the_option={
                            "是否過濾申請人": True if is_industry else False,
                            "是否過濾相同學校": True,
                            "是否過濾職稱": True,
                            "是否過濾掉自身": True
                        }
                    )
                    total_filter_result = merge_committee_advanced(total_filter_result, current_res)
                break 

            # = 3. 挑選前 10 位合格委員並寫回 Row
            qualified_members = total_filter_result["Remaining Members"]
            for i in range(1, 11):
                # 如果合格人數不足 10 位，後面會填入空值或提示
                statistical_row[f"選取委員{i}"] = qualified_members[i-1] if i <= len(qualified_members) else ""
            
            # 紀錄篩掉的原因供查核
            statistical_row["篩掉人員"] = str(total_filter_result["Filtered Members"])
            statistical_row["篩選原因"] = str(total_filter_result["Filter Reasons"])
            
            result_dict.append(statistical_row)
            
        # 儲存該分頁結果
        pd.DataFrame(result_dict).to_excel(writer, sheet_name=sheet, index=False)
    
    writer.close()
    print("過濾完成，結果已儲存。")

def load_data(file_path):
    """
        讀取 Excel 檔案並回傳 workbook 和 worksheet.
    """
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    return workbook, worksheet

def generate_letters_excel(start_index, gap, count):
    def index_to_excel_column(index):
        column = ""
        while index > 0:
            index -= 1  # 將索引調整為 0 基礎
            column = chr(index % 26 + 65) + column
            index //= 26
        return column

    letters = []
    for i in range(count):
        letter_index = start_index + i * gap
        letters.append(index_to_excel_column(letter_index))
    return letters

def add_comments(target_ws, data_ws, columns_to_comment, project_map=None, map_sheet_name=None):
    """
        在目標工作表上添加註釋。
        project_map: { "分頁|||申請計畫|||委員": 過去計畫名稱 }，
                     用於在註解中加上「此委員是依據哪一個過去計畫被推薦」。
        map_sheet_name: 查 project_map 時使用的分頁名稱（過濾後分頁名稱已改動，
                        需指定原始分頁名稱才能對應）；未指定時用 target_ws.title。
    """
    map_sheet_name = map_sheet_name or target_ws.title

    # columns_to_comment = ['D', 'F', 'H', 'J', 'L', 'N', 'P', 'R', 'T', 'V']  # 需要添加註釋的欄位

    # 建立名稱與詳細資訊的對應字典
    # ['名稱', '年份', '機關名稱', '職稱', '來源', '學校', '系所']
    name_to_details = {
    data_ws.cell(row=i, column=1).value: \
        f"名稱: {data_ws.cell(row=i, column=1).value}\n"
        f"年份: {data_ws.cell(row=i, column=2).value}\n"
        f"機關: {data_ws.cell(row=i, column=3).value}\n"
        f"職稱: {data_ws.cell(row=i, column=4).value}\n"
        # f"來源: {data_ws.cell(row=i, column=5).value}"
        for i in range(2, data_ws.max_row + 1)
    }

    headers = [cell.value for cell in data_ws[1]]  # 取得第一列的標題名稱
    source_index = headers.index('來源') + 1  # Excel 的索引從 1 開始
    # print("來源欄位的索引位置:", source_index)
    # print("欄位名稱:", headers)

    # 找出目標工作表中「計畫名稱」欄位的位置，用來查 project_map
    project_name_field = value_of_key("計畫名稱")
    target_headers = [cell.value for cell in target_ws[1]]
    proj_col_idx = target_headers.index(project_name_field) + 1 if project_name_field in target_headers else None

    # 在每個指定欄位添加註釋
    for col in columns_to_comment:
        for cell in target_ws[col]:
            if not cell.value:
                continue

            comment_parts = []
            if cell.value in name_to_details:
                comment_parts.append(name_to_details[cell.value])

            # 加上「推薦依據計畫」
            if project_map and proj_col_idx:
                apply_project = target_ws.cell(row=cell.row, column=proj_col_idx).value
                based_project = project_map.get(_make_rec_map_key(map_sheet_name, apply_project, cell.value))
                if based_project:
                    comment_parts.append(f"推薦依據計畫: {based_project}")

            if comment_parts:
                # 各段去掉尾端換行再接，避免「推薦依據計畫」與前面多空一行
                comment_text = "\n".join(part.rstrip("\n") for part in comment_parts)
                comment = openpyxl.comments.Comment(comment_text, "Python Script")
                comment.width = 350   # 設置寬度
                comment.height = 120  # 設置高度
                cell.comment = comment
                
def excel_process_VBA():
    from openpyxl.utils.cell import column_index_from_string
    from openpyxl.utils import get_column_letter

    gap = 2
    range_count = 10

    talent_workbook, talent_sheet = load_data(find_key_path("統計清單人才資料_RDF"))
    committee_workbook = openpyxl.load_workbook(find_key_path("過濾相近後統計表"))
    pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')

    run_output_dir = get_run_output_dir() or "./data/output"

    # 載入「推薦依據計畫對應表」（由 search_v3 產出）
    rec_map_path = os.path.join(run_output_dir, REC_MAP_FILENAME)
    recommend_project_map = {}
    if os.path.exists(rec_map_path):
        with open(rec_map_path, 'r', encoding='utf-8') as f:
            recommend_project_map = json.load(f)

    # 載入「過濾後遞補的合格委員名單」（由 filter_committee 產出）
    filtered_top_path = _temp_path(run_output_dir, FILTERED_TOP_FILENAME)
    filtered_top_map = {}
    if os.path.exists(filtered_top_path):
        with open(filtered_top_path, 'r', encoding='utf-8') as f:
            filtered_top_map = json.load(f)

    project_name_field = value_of_key("計畫名稱")
    no_fill = PatternFill(fill_type=None)

    # 先取得原始分頁清單（稍後會新增「_過濾後」分頁，避免邊迭代邊新增）
    original_sheets = list(committee_workbook.worksheets)

    for committee_sheet in original_sheets:
        out_count    = committee_sheet.max_column - (range_count * gap) - 6
        start_index  = out_count + 1
        manager_cols = generate_letters_excel(start_index, gap, range_count)        # 推薦委員 1..10
        score_cols   = generate_letters_excel(start_index + 1, gap, range_count)    # 相關分數 1..10

        # 找出此分頁「計畫名稱」欄位位置（給過濾後分頁查 backfill 用）
        target_headers = [c.value for c in committee_sheet[1]]
        proj_col_idx = target_headers.index(project_name_field) + 1 if project_name_field in target_headers else None

        # ===== 原始分頁：加註解、被篩掉者標粉紅 =====
        add_comments(committee_sheet, talent_sheet, columns_to_comment=manager_cols, project_map=recommend_project_map)

        header_value = committee_sheet.cell(row=1, column=start_index).value
        print("[{}] 第 {} 欄之標題為：{}".format(committee_sheet.title, start_index, header_value))

        for row in committee_sheet.iter_rows(min_row=2, max_col=committee_sheet.max_column):
            filter_list = ast.literal_eval(row[-2].value)
            for col_letter in manager_cols:
                col_index = column_index_from_string(col_letter) - 1
                if row[col_index].value in filter_list:
                    row[col_index].fill = pink_fill

        # ===== 過濾後分頁：移除被篩掉者、由候選池遞補到 10 位 =====
        filtered_sheet = committee_workbook.copy_worksheet(committee_sheet)
        filtered_sheet.title = (committee_sheet.title + "_過濾後")[:31]

        for row in filtered_sheet.iter_rows(min_row=2, max_col=filtered_sheet.max_column):
            apply_project = row[proj_col_idx - 1].value if proj_col_idx else None
            backfilled = filtered_top_map.get(_make_pool_key(committee_sheet.title, apply_project), [])

            for k, (m_letter, s_letter) in enumerate(zip(manager_cols, score_cols)):
                m_idx = column_index_from_string(m_letter) - 1
                s_idx = column_index_from_string(s_letter) - 1
                if k < len(backfilled):
                    row[m_idx].value = backfilled[k]['manager']
                    row[s_idx].value = backfilled[k]['score']
                else:
                    row[m_idx].value = None
                    row[s_idx].value = None
                row[m_idx].fill = no_fill       # 清掉由原分頁複製來的粉紅標記
                row[m_idx].comment = None       # 清掉複製來的舊註解，稍後重新加上

        # 過濾後分頁重新加註解（用原始分頁名稱查推薦依據計畫）
        add_comments(filtered_sheet, talent_sheet, columns_to_comment=manager_cols,
                     project_map=recommend_project_map, map_sheet_name=committee_sheet.title)

        # 讓「_過濾後」分頁緊接在對應原始分頁之後
        sheets = committee_workbook._sheets
        sheets.remove(filtered_sheet)
        sheets.insert(sheets.index(committee_sheet) + 1, filtered_sheet)

    committee_workbook.save(find_key_path("FINAL_COMMITTEE"))
    print(f"[結束] 已經保存至: {find_key_path('FINAL_COMMITTEE')}")

from datetime import datetime
def update_peronsal_info_database(is_industry=False):
    '''
        更新暫存最新人才資料庫
    '''
    
    # 讀取最新 Excel 檔案
    if is_industry:
        apply_list_folder_path = find_key_path("產學合作申請名冊")
        file_name = value_of_key("產學合作申請名冊")
    else:
        apply_list_folder_path = find_key_path("研究計畫申請名冊") 
        file_name = value_of_key("研究計畫申請名冊")  # 修正錯誤的 `file_name(file_name)`

    apply_list_file = pd.ExcelFile(apply_list_folder_path)
    
    # 讀取暫存人才資料庫，若不存在則建立新的 DataFrame
    personal_info_database_path = find_key_path("暫存最新人才資料庫")

    if os.path.exists(personal_info_database_path):
        personal_info_database = pd.read_excel(personal_info_database_path)
    else:
        personal_info_database = pd.DataFrame(columns=['名稱', '年份', '機關名稱', '職稱', '來源'])

    new_data = []

    for sheet in apply_list_file.sheet_names: 
        current_sheet_apply_excel_data = pd.read_excel(apply_list_file, sheet_name=sheet)
        for _, row in current_sheet_apply_excel_data.iterrows():
            new_row = {
                '名稱': row.get(value_of_key('申請主持人欄位名稱')), 
                '年份': datetime.now().year - 1911,  # 直接填入現在的年份
                '機關名稱': row.get(value_of_key('申請機構欄位名稱'), ''),
                '職稱': row.get(value_of_key('職稱'), ''),
                '來源': file_name,
            }
            new_data.append(new_row)

    # 將新數據轉換為 DataFrame
    new_data_df = pd.DataFrame(new_data)

    # 合併數據，並以 "名稱" 為主鍵，保留最新年份的資料
    updated_database = pd.concat([personal_info_database, new_data_df]).sort_values(by=['名稱', '年份'], ascending=[True, False])
    updated_database = updated_database.drop_duplicates(subset=['名稱'], keep='first')  # 只保留最新的年份

    # 儲存回 Excel
    updated_database.to_excel(personal_info_database_path, index=False)